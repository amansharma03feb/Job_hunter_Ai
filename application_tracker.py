"""
Application Tracker — logs every outreach to an Excel file.

Columns include JD summary, ATS score breakdown, HR details, Drive links,
and application status — per Rule 4.
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LOG_FILE = "Job_Hunt_Tracker.xlsx"

HEADERS = [
    "Date",
    "Company",
    "Role",
    "Location",
    "JD Summary",
    "JD Source URL",
    "ATS Score %",
    "Keyword Score %",
    "AI Score %",
    "AI Verdict",
    "HR Name",
    "HR Title",
    "HR Email",
    "Email Confidence",
    "Email Status",
    "Resume Version",
    "Resume (Drive)",
    "Cover Letter (Drive)",
    "Apply URL",
    "Status",
    "Easy Apply",
    "Direct Applied",
    "Outreach Touch",
    "Last Touch Date",
    "Next Action",
]

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")

COL_WIDTHS = [
    14,  # Date
    22,  # Company
    35,  # Role
    18,  # Location
    50,  # JD Summary
    40,  # JD Source URL
    12,  # ATS Score %
    12,  # Keyword Score %
    12,  # AI Score %
    14,  # AI Verdict
    22,  # HR Name
    30,  # HR Title
    36,  # HR Email
    16,  # Email Confidence
    14,  # Email Status
    20,  # Resume Version
    45,  # Resume (Drive)
    45,  # Cover Letter (Drive)
    55,  # Apply URL
    14,  # Status
    10,  # Easy Apply
    12,  # Direct Applied
    14,  # Outreach Touch
    14,  # Last Touch Date
    18,  # Next Action
]


def _init_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    return wb, ws


def _compute_next_action(status, touch, confidence):
    """Determine next action based on current state."""
    if status in ("Failed", "No Contact", "Duplicate"):
        return ""
    if confidence in ("generic_alias", "hr_alias"):
        return "Find real contact"
    if touch >= 3:
        return "Move on"
    if touch == 2:
        return "Final follow-up Day 10"
    return "Follow-up Day 5"


def _truncate_jd(description, max_words=100):
    """Truncate JD to max_words for the summary column."""
    if not description:
        return ""
    words = str(description).split()
    if len(words) <= max_words:
        return " ".join(words)
    return " ".join(words[:max_words]) + "..."


def log_application(job, hr_info, hr_email, email_confidence, email_status,
                    resume_file="", resume_drive_link="", cl_drive_link=""):
    """
    Append one row to the application log Excel file.
    Creates the file with new headers if it doesn't exist.
    """
    os.makedirs("output", exist_ok=True)

    if os.path.exists(LOG_FILE):
        wb = load_workbook(LOG_FILE)
        ws = wb.active
        current_headers = [ws.cell(row=1, column=c).value for c in range(1, len(HEADERS) + 1)]
        if current_headers != HEADERS:
            backup = LOG_FILE.replace(".xlsx", "_backup.xlsx")
            wb.save(backup)
            wb.close()
            wb, ws = _init_workbook()
            print(f"   [Tracker] Old format backed up to {backup}, new tracker created")
    else:
        wb, ws = _init_workbook()

    next_row = ws.max_row + 1
    use_alt  = (next_row % 2 == 0)

    # Determine status
    if email_status == "sent":
        status = "Applied"
    elif email_status == "already contacted":
        status = "Duplicate"
    elif "failed" in str(email_status):
        status = "Failed"
    elif email_status == "email_not_found":
        status = "No Contact"
    else:
        status = "Drafted"

    # Detect Easy Apply and direct apply URL
    easy_apply = "Yes" if job.get("easy_apply") else ""
    apply_url = job.get("apply_url") or job.get("url", "")
    direct_applied = "Yes" if job.get("_direct_applied") else ""

    # Outreach touch tracking
    touch = job.get("_outreach_touch", 1)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    next_action = _compute_next_action(status, touch, email_confidence)

    values = [
        now_str,                                                          # Date
        job.get("company", ""),                                           # Company
        job.get("title", ""),                                             # Role
        job.get("location", ""),                                          # Location
        _truncate_jd(job.get("description", "")),                         # JD Summary
        job.get("url", ""),                                               # JD Source URL
        job.get("final_score", ""),                                       # ATS Score %
        job.get("keyword_score", ""),                                     # Keyword Score %
        job.get("ai_score", "") if job.get("ai_score") is not None else "",  # AI Score %
        job.get("ai_verdict", ""),                                        # AI Verdict
        hr_info.get("full_name", ""),                                     # HR Name
        hr_info.get("title", ""),                                         # HR Title
        hr_email or "",                                                   # HR Email
        email_confidence or "",                                           # Email Confidence
        email_status,                                                     # Email Status
        os.path.basename(resume_file) if resume_file else "master_resume.pdf",  # Resume Version
        resume_drive_link or "",                                          # Resume (Drive)
        cl_drive_link or "",                                              # Cover Letter (Drive)
        apply_url,                                                        # Apply URL
        status,                                                           # Status
        easy_apply,                                                       # Easy Apply
        direct_applied,                                                   # Direct Applied
        f"Touch {touch}",                                                 # Outreach Touch
        now_str[:10],                                                     # Last Touch Date
        next_action,                                                      # Next Action
    ]

    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.alignment = Alignment(wrap_text=(col == 5), vertical="center")  # Wrap JD summary
        if use_alt:
            cell.fill = _ALT_FILL

    # Hyperlinks on URL columns
    url_columns = {6: job.get("url", ""), 19: job.get("url", "")}
    for col_idx, url in url_columns.items():
        if url:
            ws.cell(row=next_row, column=col_idx).hyperlink = url
            ws.cell(row=next_row, column=col_idx).font = Font(color="0563C1", underline="single")

    # Hyperlinks on Drive columns (17, 18)
    if resume_drive_link:
        ws.cell(row=next_row, column=17).hyperlink = resume_drive_link
        ws.cell(row=next_row, column=17).font = Font(color="0563C1", underline="single")
    if cl_drive_link:
        ws.cell(row=next_row, column=18).hyperlink = cl_drive_link
        ws.cell(row=next_row, column=18).font = Font(color="0563C1", underline="single")

    wb.save(LOG_FILE)
    return LOG_FILE


def get_already_contacted():
    """Return set of company+role pairs already emailed (avoid duplicates)."""
    if not os.path.exists(LOG_FILE):
        return set()
    try:
        wb = load_workbook(LOG_FILE, read_only=True)
        ws = wb.active
        seen = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            company = str(row[1] or "").lower()
            role    = str(row[2] or "").lower()
            seen.add(f"{company}|{role}")
        wb.close()
        return seen
    except Exception:
        return set()


def get_followup_candidates(days_threshold=5):
    """
    Find companies needing follow-up: contacted N+ days ago, status Applied/Followed Up,
    personal email (not generic), and haven't hit 3 touches yet.
    Returns list of dicts.
    """
    if not os.path.exists(LOG_FILE):
        return []
    try:
        wb = load_workbook(LOG_FILE, read_only=True)
        ws = wb.active
        candidates = []
        now = datetime.now()
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Safely get columns — handle both old (20-col) and new (25-col) formats
            cols = list(row) + [None] * 10  # pad to avoid IndexError
            company    = str(cols[1] or "")
            role       = str(cols[2] or "")
            hr_name    = str(cols[10] or "")
            hr_email   = str(cols[12] or "")
            confidence = str(cols[13] or "")
            status     = str(cols[19] or "")
            touch_str  = str(cols[22] or "Touch 1")
            last_date  = str(cols[23] or cols[0] or "")

            if status not in ("Applied", "Followed Up") or not hr_email:
                continue
            if confidence in ("generic_alias", "hr_alias"):
                continue

            # Parse touch number
            try:
                touch_num = int(touch_str.replace("Touch ", ""))
            except (ValueError, AttributeError):
                touch_num = 1
            if touch_num >= 3:
                continue

            try:
                check_date = datetime.strptime(str(last_date)[:10], "%Y-%m-%d")
            except (ValueError, IndexError):
                continue

            days_ago = (now - check_date).days
            if days_ago >= days_threshold:
                candidates.append({
                    "company": company,
                    "role": role,
                    "hr_email": hr_email,
                    "hr_name": hr_name,
                    "days_ago": days_ago,
                    "touch": touch_num,
                })
        wb.close()
        return candidates
    except Exception:
        return []


def update_status(company, role, new_status, increment_touch=False):
    """Update Status, touch count, and last touch date for a company+role entry."""
    if not os.path.exists(LOG_FILE):
        return
    try:
        wb = load_workbook(LOG_FILE)
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            c = str(ws.cell(row=row_idx, column=2).value or "").lower()
            r = str(ws.cell(row=row_idx, column=3).value or "").lower()
            if c == company.lower() and r == role.lower():
                ws.cell(row=row_idx, column=20, value=new_status)
                if increment_touch:
                    old_touch = str(ws.cell(row=row_idx, column=23).value or "Touch 1")
                    try:
                        num = int(old_touch.replace("Touch ", "")) + 1
                    except ValueError:
                        num = 2
                    ws.cell(row=row_idx, column=23, value=f"Touch {num}")
                    ws.cell(row=row_idx, column=24, value=datetime.now().strftime("%Y-%m-%d"))
                    ws.cell(row=row_idx, column=25, value=_compute_next_action(new_status, num, ""))
                break
        wb.save(LOG_FILE)
    except Exception:
        pass
